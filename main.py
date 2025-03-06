import cv2
import time
import threading
from datetime import datetime
from RPLCD.i2c import CharLCD
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import gspread
from google.oauth2.service_account import Credentials
import face_recognition
import cv2
import json
import os
import glob
import numpy as np
import requests
import re


cap = cv2.VideoCapture(0)
lcd=CharLCD(i2c_expander='PCF8574', address=0x27, port=1, cols=32, rows=2, dotsize=8)
lcd.clear()
ultimaDeteccion = None
ultimoPrint = None
known_face_encodings = []
known_face_names = []
contadores_deteccion = {}
umbral=4
frame_resizing = 0.25
carpeta="conocidos/"
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
creds = Credentials.from_service_account_file('speedy-solstice-434503-m6-161d71670782.json', scopes=SCOPES)
client = gspread.authorize(creds)
url = "http://localhost:3008/v1/messages"
inaFill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
inaFont= Font(name="Calibri", size=11, bold=True, color="9c0006")
retFill=PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid")
retFont=Font(name="Calibri", size=11, bold=True, color="815c00")

alumnosDes = [
    os.path.splitext(f)[0]
    for f in os.listdir(carpeta)
    if os.path.isfile(os.path.join(carpeta, f)) and f.lower().endswith(('.jpeg', '.jpg', '.png'))
]

alumnos=sorted(alumnosDes)

modulos=[[0,1,2,3,4,5,6,7,8],
        ["0","7.00-7.50","7.50-8.40","8.40-9.20","9.20-10.05","10.05-10.50","10.50-11.40","11.40-12.30","12.30-13.20"]]

modulosRangos= [
        (420, 470, 1),  #7:00-7:50 (420-470)
        (470, 520, 2),  #7:50-8:40 (470-520)
        (520, 560, 3),  #8:40-9:20 (520-560)
        (560, 605, 4),  #9:20-10:05 (560-605)
        (605, 650, 5),  #10:05-10:50 (605-650)
        (650, 700, 6),  #10:50-11:40 (650-700)
        (700, 750, 7),  #11:40-12:30 (700-750)
        (750, 800, 8),  #12:30-13:20 (750-800)
    ]

matrizMaterias=[]

matrizTutores=[]

modulos_procesados = set()

dias=["LUNES","MARTES","MIERCOLES","JUEVES","VIERNES"]

def guardarDatos():
        with open("cacheM.json","w") as f:
                json.dump(matrizMaterias,f,indent=4)
                
        with open("cacheT.json","w") as f:
                json.dump(matrizTutores,f,indent=4)        
                
def cargarDatos():
        global matrizMaterias,matrizTutores
        with open("cacheM.json","r") as f:
                matrizMaterias=json.load(f)
                
        with open("cacheT.json","r") as f:
                matrizTutores=json.load(f)   

def obtenerDatos():
    global matrizMaterias, matrizTutores 
    try:
        horarioID = '1GFUgzdS-nKockbae-v5upwNBLaYEY6LhwVSiLAlfrIw'
        spreadsheet = client.open_by_key(horarioID)
    
    
        sheet = spreadsheet.worksheet(dias[datetime.now().weekday()])
        data = sheet.get_all_records() 
        matrizMaterias = [row for row in data]
        
        
        tutorID = '1GrPbYbjPwrsLT6d4Y98UyI9pMo0vIg6GR6T6OsxJOto'
        spreadsheet = client.open_by_key(tutorID)
    
    
        sheet = spreadsheet.worksheet("6AMP")
        data = sheet.get_all_records() 
        matrizTutores = [row for row in data]
        
        guardarDatos()
        
    except Exception as e:
        print(f"Ocurrió un error inesperado: {e}")
        return "(No detectado)"


def buscarTutor(name):
    for y in range(len(matrizTutores)):
        if matrizTutores[y]["ALUMNO"].lower() == name.lower():
            if matrizTutores[y]["TUTOR"] != "":

                return matrizTutores[y]["TUTOR"], matrizTutores[y]["NUMERO"]
    return "",""

def print2(mensaje):
    global ultimoPrint
    if mensaje != ultimoPrint:
        print(mensaje)
        ultimoPrint=mensaje

def get_module_by_time(entry_time):
    global modulosRangos
    hour = entry_time.hour
    minute = entry_time.minute
    time_in_minutes = hour * 60 + minute  


    for start, end, module_index in modulosRangos:
        if start <= time_in_minutes < end:
            return f'Modulo {module_index} ({modulos[1][module_index]})'

    return ""

def plantilla(ws):
    fecha = datetime.now().strftime("%d/%m/%Y")
    
    titulo_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    encabezado_font = Font(name='Calibri', size=12, bold=True)
    contenido_font = Font(name='Calibri', size=11)
    
   
    tinto = PatternFill(start_color='8b0101', end_color='8b0101', fill_type='solid')
    rojo = PatternFill(start_color='bf4646', end_color='bf4646', fill_type='solid')
    
    
    borde_completo = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:D1')
    ws['A1'] = "Centro de Estudios Tecnologico Industrial y de Servicios no.107"
    ws['A1'].font = titulo_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = tinto
    

    ws['A2'] = f'Grupo: 6AMP'
    ws['A2'].font = encabezado_font
    ws['D2'] = f'Fecha: {fecha}'
    ws['D2'].font = encabezado_font
    
    encabezados = ['No.', 'Nombre', 'Asistencia', 'Observaciones']
    for col, encabezado in enumerate(encabezados, start=1):
        cell = ws.cell(row=3, column=col, value=encabezado)
        cell.font = encabezado_font
        cell.fill = rojo
        cell.alignment = Alignment(horizontal='center')
        cell.border = borde_completo
    
    for index, alumno in enumerate(alumnos, start=1):
        ws.cell(row=index+4, column=1, value=index).font = contenido_font
        ws.cell(row=index+4, column=2, value=alumno).font = contenido_font
        
 
        for col in [3, 4]:
            cell = ws.cell(row=index+4, column=col)
            cell.font = contenido_font
            cell.border = borde_completo    
  

    ws.column_dimensions['A'].width = 5   # Número
    ws.column_dimensions['B'].width = 45  # Nombre
    ws.column_dimensions['C'].width = 15  # Asistencia
    ws.column_dimensions['D'].width = 35  # Observaciones

def excel():
    global modulos, modulosRangos, inaFill, inaFont, retFill, retFont
    date_str = datetime.now().strftime('%Y-%m-%d')  
    file_name = f"semestre/{date_str}.xlsx"
    
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        """for sheet in wb.sheetnames:
            plantilla(wb[sheet]) (?)"""
    else:
        wb = Workbook()
        wb.remove(wb.active)
        for i in range(1, len(modulos[0])):
            sheet_name2 = f'Modulo {i} (' + modulos[1][i] + ")"
            ws = wb.create_sheet(sheet_name2)
            plantilla(ws)
    wb.save(file_name) 

def add_to_excel(name, entry_time):
    global modulos, modulosRangos, inaFill, inaFont, retFill, retFont
    contador=1
    entry_time = datetime.fromtimestamp(entry_time)
    hour = entry_time.hour
    minute = entry_time.minute
    time_in_minutes = hour * 60 + minute 
    date_str = datetime.now().strftime('%Y-%m-%d')  
    entry_time_str = entry_time.strftime('%H:%M')
    file_name = f"{date_str}.xlsx"
    
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        """for sheet in wb.sheetnames:
            plantilla(wb[sheet]) (?)"""
    else:
        wb = Workbook()
        wb.remove(wb.active)
        for i in range(1, len(modulos[0])):
            sheet_name2 = f'Modulo {i} (' + modulos[1][i] + ")"
            ws = wb.create_sheet(sheet_name2)
            plantilla(ws)


    module_sheet = get_module_by_time(entry_time)
    moduloActual= int(re.search(r'\d+', module_sheet).group()) if re.search(r'\d+', module_sheet) else 0


    if module_sheet != "":
        sheet = wb[module_sheet]
        fila = max(sheet.max_row + 1, 4)
        if sheet.max_row == 3: 
            sheet.cell(row=fila, column=1).value = contador  
            sheet.cell(row=fila, column=2).value = name
            sheet.cell(row=fila, column=3).value = entry_time_str

        else:
            registered_names = {row[0] for row in sheet.iter_rows(min_row=4, min_col=2, max_col=2, values_only=True) if row[0]}
            celda = sheet.cell(row=alumnos.index(name) + 5, column=3)
            if name in registered_names and not celda.value:
                print2(name+" "+entry_time_str)
                sheet.cell(row=alumnos.index(name) + 5, column=3).value = entry_time_str

                
                if modulosRangos[moduloActual-1][0]+5 >= time_in_minutes:
                    print(name+" registrad@ a tiempo a las "+entry_time_str)
                    
                elif modulosRangos[moduloActual-1][0]+49 >= time_in_minutes:
                    sheet.cell(row=alumnos.index(name) + 5, column=3).fill = retFill
                    sheet.cell(row=alumnos.index(name) + 5, column=3).font = retFont
                    print(name+" registrad@ en retardo a las "+entry_time_str+" (más de 5 minutos tarde)")
                    
                elif modulosRangos[moduloActual-1][0]+49 < time_in_minutes:
                    sheet.cell(row=alumnos.index(name) + 5, column=3).fill = inaFill
                    sheet.cell(row=alumnos.index(name) + 5, column=3).font = inaFont
                    print(name+" registrad@ como inasistencia a las "+entry_time_str+" (más de 10 minutos tarde)")            
     
        if time_in_minutes > modulosRangos[moduloActual - 1][1]  :
            for row in range(4, sheet.max_row + 1):  
                celda = sheet.cell(row=row, column=3)  
                if not celda.value:  
                    celda.fill = inaFill      
                                      
        wb.save(file_name)  
                
    else:        
        print2("Se detecto a "+name+", más no hay ningún modulo establecido para este horario ("+entry_time_str+")")  
   
        
def load_encoding_images(images_path):

    images_path = glob.glob(os.path.join(images_path, "*.*"))

    print("Se han encontrado {} imagenes en la base de datos. Codificando...".format(len(images_path)))
    lcd.clear()
    lcd.crlf()
    lcd.write_string("Codificando...")


    for img_path in images_path:
        img = cv2.imread(img_path)
        rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        basename = os.path.basename(img_path)
        (filename, ext) = os.path.splitext(basename)
        img_encoding = face_recognition.face_encodings(rgb_img)[0]
        known_face_encodings.append(img_encoding)
        known_face_names.append(filename)
    print("Imagenes cargadas. En ejecución")
    lcd.clear()
    lcd.crlf()
    lcd.write_string("En ejecucion")

def detect_known_faces(frame):
    global ultimaDeteccion, contadores_deteccion
    
    small_frame = cv2.resize(frame, (0, 0), fx=frame_resizing, fy=frame_resizing)
    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
    

    face_locations = face_recognition.face_locations(rgb_small_frame)
    face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

    face_names = []
    
    for face_encoding in face_encodings:
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=0.495)
        name = ""

        face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
        best_match_index = np.argmin(face_distances)
        
        if matches[best_match_index]:
            name = known_face_names[best_match_index]
            print(f"Se cree detectar a {name}")
            
            
            if name in contadores_deteccion:
                contadores_deteccion[name]+=1
            else:
                contadores_deteccion[name]=1
            if contadores_deteccion[name]>=umbral:
                if ultimaDeteccion!=name:
                    print2(f"Se detecto a  {name}.")
                    lcd.clear()
                    lcd.write_string(f"Bienvenido, ")
                    lcd.crlf()
                    lcd.write_string(re.search(r'^\S+', name).group())
                    add_to_excel(name, datetime.now().timestamp())
                    ultimaDeteccion=name
        else:
             if name in contadores_deteccion and ultimaDeteccion!= name:
                 contadores_deteccion[name]=0                    

        face_names.append(name)

    if len(face_locations) == 0:
        return [], []

    face_locations = np.array(face_locations)
    face_locations = (face_locations / frame_resizing).astype(int)
    
    return face_locations, face_names

def marcar_absences(modulo):
    date_str = datetime.now().strftime('%Y-%m-%d')  
    file_name = f"{date_str}.xlsx"

    if not os.path.exists(file_name):
        print(f"No hay archivo de asistencia para hoy ({file_name}).")
        return
    
    wb = load_workbook(file_name)
    sheet_name = f'Modulo {modulo} (' + modulos[1][modulo] + ")"
    
    if sheet_name not in wb.sheetnames:
        print(f"No hay hoja para el módulo {modulo}.")
        return
    
    ws = wb[sheet_name]
    
    for row in range(5, ws.max_row + 1):
        cell_name = ws.cell(row=row, column=2).value
        cell_asistencia = ws.cell(row=row, column=3)
        
        if (cell_name and cell_asistencia.value is None) or cell_name and cell_asistencia.fill==inaFill:
            print(cell_name) 
            cell_asistencia.value = "-"
            cell_asistencia.fill = inaFill
            cell_asistencia.font = inaFont
            tutor,numero=buscarTutor(cell_name)
            if tutor !="" or numero !="":
                response = requests.post(url, json={"number": "521"+str(numero), "message": f"Estimad@ {tutor}, su hij@ {cell_name} ha sido registrad@ con una inasistencia en el modulo {modulo}"})
                if response.status_code == 200:
                    print("Mensaje enviado con éxito")
                else:
                    print("Error al enviar mensaje:", response.text)
    wb.save(file_name)
    print(f"Inasistencias marcadas en el módulo {modulo}.")

def verificar_tiempo():
    while True:
        hora_actual = datetime.now().hour * 60 + datetime.now().minute  

        for inicio, fin, modulo in modulosRangos:
            if hora_actual >= fin and modulo not in modulos_procesados:
                marcar_absences(modulo) 
                modulos_procesados.add(modulo) 
                
        time.sleep(30) 


obtenerDatos()
print(matrizMaterias)
print(matrizTutores)

if len(matrizMaterias)==0:
        print("No se pudo acceder a google sheets. Cargando datos mas recientes...")
        cargarDatos()
else:print("Conectado a google sheets. Cargando datos...")

load_encoding_images("conocidos/")
threading.Thread(target=verificar_tiempo, daemon=True).start()


try:
 while True:
    ret, frame = cap.read()
    face_locations, face_names = detect_known_faces(frame)
    for face_loc, name in zip(face_locations, face_names):
        y1, x2, y2, x1 = face_loc[0], face_loc[1], face_loc[2], face_loc[3]
    
        actual = time.time()
        
        if  name !="":
            cv2.putText(frame, "Bienvenid@, " + name, (20, 450), cv2.FONT_HERSHEY_COMPLEX_SMALL, 1, (3, 192, 60) if name != "" else (195, 59, 35), 2)
        color = (3, 192, 60) if name != "" else (255,255,255)
        cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)

    cv2.imshow("GRAN HERMANO", frame)

    key = cv2.waitKey(1)
    if key == ord("q"):
        lcd.clear()
        lcd.write_string('PUNTO DE CONTROL   ')
        lcd.crlf()
        lcd.write_string("AT4580")
        break
        
     

 cap.release()
 cv2.destroyAllWindows()


except Exception as e:
    lcd.write_string('PUNTO DE CONTROL   ')
    lcd.crlf()
    lcd.write_string("AT4580")
